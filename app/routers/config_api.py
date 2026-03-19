"""
routers/config_api.py — Auditor IPs
Settings, backup/restore BD, push notifications, VAPID,
TLS info, export XLSX y test Discord.

Scheduler se inyecta desde main.py vía set_scheduler().
"""

import base64
import csv
import io
import os
import shutil
import sqlite3
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

from fastapi import APIRouter, Body, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse, Response

from config import cfg, cfg_defaults, save_setting, DB_PATH, SCAN_CIDR, SCAN_INTERVAL_SECONDS
from database import db
from utils import utc_now_iso, to_local_str, human_since, get_app_tz

router = APIRouter()

_scheduler_ref: Any = None


def set_scheduler(sched: Any) -> None:
    global _scheduler_ref
    _scheduler_ref = sched
    _reschedule_export()          # arrancar job de export si está configurado


def _get_scheduler():
    return _scheduler_ref


# ═══════════════════════════════════════════════════════════════
#  Settings
# ═══════════════════════════════════════════════════════════════

@router.get("/api/settings")
def api_get_settings():
    from config import _cfg
    return {"ok": True, "settings": dict(_cfg)}


@router.put("/api/settings")
def api_put_settings(payload: Dict[str, Any] = Body(...)):
    allowed = set(cfg_defaults().keys()) | {
        "primary_net_label", "hidden_tabs", "ui_lang", "wol_public",
        "smtp_enabled", "smtp_host", "smtp_port", "smtp_tls",
        "smtp_user", "smtp_pass", "smtp_to", "smtp_from",
    }  # extra keys not in defaults
    updated = []
    for k, v in payload.items():
        if k not in allowed:
            continue
        save_setting(k, str(v))
        updated.append(k)

    # Reagendar scan si cambió intervalo o CIDR
    if "scan_interval" in updated or "scan_cidr" in updated:
        new_interval = int(cfg("scan_interval", SCAN_INTERVAL_SECONDS))
        sched = _get_scheduler()
        if sched:
            try:
                sched.reschedule_job("scan_job", trigger="interval", seconds=new_interval)
            except Exception:
                from routers.scans import run_scan
                sched.add_job(
                    lambda: run_scan(cfg("scan_cidr", SCAN_CIDR)),
                    "interval", seconds=new_interval,
                    id="scan_job", replace_existing=True,
                )

    return {"ok": True, "updated": updated}


@router.post("/api/settings/test-discord")
def api_test_discord():
    webhook = cfg("discord_webhook", "")
    if not webhook:
        return JSONResponse({"ok": False, "error": "No hay webhook configurado"}, status_code=400)
    from routers.scans import discord_notify
    ok, err = discord_notify("🧪 **Test Auditor IPs** — conexión Discord funcionando ✅")
    return {"ok": ok, "error": err}


# ── SMTP email ────────────────────────────────────────────────

def send_email(subject: str, body: str) -> tuple[bool, str]:
    """
    Envía un email usando la configuración SMTP almacenada en BD.
    Devuelve (ok, error_msg).
    Usa solo stdlib: smtplib + email.
    """
    import smtplib
    import ssl
    from email.message import EmailMessage

    enabled = cfg("smtp_enabled", "0")
    if enabled != "1":
        return False, "SMTP no habilitado"

    host     = cfg("smtp_host",  "").strip()
    port     = int(cfg("smtp_port", "587") or 587)
    tls_mode = cfg("smtp_tls",   "starttls")
    user     = cfg("smtp_user",  "").strip()
    password = cfg("smtp_pass",  "")
    to_addr  = cfg("smtp_to",    "").strip()
    from_addr= cfg("smtp_from",  "").strip() or user

    if not host or not to_addr:
        return False, "Faltan host SMTP o dirección destinatario"

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"]    = from_addr
    msg["To"]      = to_addr
    msg.set_content(body)

    try:
        if tls_mode == "ssl":
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=15) as s:
                if user and password:
                    s.login(user, password)
                s.send_message(msg)
        elif tls_mode == "starttls":
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.ehlo()
                s.starttls(context=ssl.create_default_context())
                s.ehlo()
                if user and password:
                    s.login(user, password)
                s.send_message(msg)
        else:  # none
            with smtplib.SMTP(host, port, timeout=15) as s:
                if user and password:
                    s.login(user, password)
                s.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)


@router.post("/api/settings/test-smtp")
def api_test_smtp():
    """Envía un email de prueba con la configuración SMTP actual."""
    ok, err = send_email(
        subject="🧪 Test Auditor IPs — SMTP funcionando",
        body="Este es un mensaje de prueba del sistema de alertas de Auditor IPs.\n\nSi recibes este correo, la configuración SMTP es correcta."
    )
    return {"ok": ok, "error": err or None}


# ═══════════════════════════════════════════════════════════════
#  Backup / Restore
# ═══════════════════════════════════════════════════════════════

BACKUP_DIR = os.path.join(os.path.dirname(DB_PATH), "backups")


def run_backup() -> Dict[str, Any]:
    """Crea copia de seguridad de la BD con la SQLite backup API."""
    if cfg("backup_enabled", "1") != "1":
        return {"ok": True, "skipped": True}
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(BACKUP_DIR, f"auditor_{ts}.db")
        src_conn = sqlite3.connect(DB_PATH)
        dst_conn = sqlite3.connect(dest)
        src_conn.backup(dst_conn)
        src_conn.close()
        dst_conn.close()
        size_kb = round(os.path.getsize(dest) / 1024, 1)
        keep    = int(cfg("backup_keep", "7"))
        all_bak = sorted(
            [f for f in os.listdir(BACKUP_DIR) if f.startswith("auditor_") and f.endswith(".db")],
            reverse=True,
        )
        removed = []
        for old in all_bak[keep:]:
            os.remove(os.path.join(BACKUP_DIR, old))
            removed.append(old)
        return {"ok": True, "file": dest, "size_kb": size_kb, "removed": removed}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.post("/api/backup/run")
def api_backup_run():
    return run_backup()


@router.get("/api/backup/list")
def api_backup_list():
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        files = sorted(
            [f for f in os.listdir(BACKUP_DIR) if f.startswith("auditor_") and f.endswith(".db")],
            reverse=True,
        )
        backups = []
        for f in files:
            path = os.path.join(BACKUP_DIR, f)
            stat = os.stat(path)
            backups.append({
                "filename": f,
                "size_kb":  round(stat.st_size / 1024, 1),
                "created":  datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
        return {"ok": True, "backups": backups, "backup_dir": BACKUP_DIR}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.get("/api/backup/download/{filename}")
def api_backup_download(filename: str):
    if "/" in filename or "\\" in filename or not filename.endswith(".db"):
        return JSONResponse({"ok": False, "error": "Nombre inválido"}, status_code=400)
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.isfile(path):
        return JSONResponse({"ok": False, "error": "Archivo no encontrado"}, status_code=404)
    return FileResponse(path, filename=filename, media_type="application/octet-stream")


@router.delete("/api/backup/{filename}")
def api_backup_delete(filename: str):
    if "/" in filename or "\\" in filename or not filename.endswith(".db"):
        return JSONResponse({"ok": False, "error": "Nombre inválido"}, status_code=400)
    path = os.path.join(BACKUP_DIR, filename)
    if not os.path.isfile(path):
        return JSONResponse({"ok": False, "error": "Archivo no encontrado"}, status_code=404)
    os.remove(path)
    return {"ok": True}


@router.get("/api/db/backup")
def db_backup_download():
    """Descarga el fichero .db completo en vivo."""
    app_tz   = get_app_tz(cfg("app_tz", "Europe/Madrid"))
    tmp      = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    tmp.close()
    shutil.copy2(DB_PATH, tmp.name)
    filename = f"auditor_ips_backup_{datetime.now(app_tz).strftime('%Y%m%d_%H%M%S')}.db"
    return FileResponse(
        tmp.name, media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        filename=filename,
    )


@router.post("/api/db/restore")
async def db_restore(request: Request):
    """Restaura la BD desde un fichero .db subido por formulario."""
    from fastapi import UploadFile
    form = await request.form()
    file: UploadFile = form.get("file")
    if not file:
        return JSONResponse({"ok": False, "error": "No se recibió fichero"}, status_code=400)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    tmp.close()
    try:
        content = await file.read()
        if not content.startswith(b"SQLite format 3"):
            return JSONResponse({"ok": False, "error": "No es una BD SQLite válida"}, status_code=400)
        with open(tmp.name, "wb") as f:
            f.write(content)
        shutil.copy2(tmp.name, DB_PATH)
        return {"ok": True, "message": "BD restaurada. Recarga la página."}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ═══════════════════════════════════════════════════════════════
#  Push notifications
# ═══════════════════════════════════════════════════════════════

@router.post("/api/push/subscribe")
def api_push_subscribe(payload: Dict[str, Any] = Body(...)):
    endpoint = (payload.get("endpoint") or "").strip()
    p256dh   = (payload.get("p256dh")   or "").strip()
    auth     = (payload.get("auth")     or "").strip()
    if not endpoint or not p256dh or not auth:
        return JSONResponse({"ok": False, "error": "Faltan campos"}, status_code=400)
    with db() as conn:
        conn.execute("""
            INSERT OR REPLACE INTO push_subscriptions (endpoint, p256dh, auth, created_at)
            VALUES (?,?,?,?)
        """, (endpoint, p256dh, auth, utc_now_iso()))
    return {"ok": True}


@router.post("/api/push/unsubscribe")
def api_push_unsubscribe(payload: Dict[str, Any] = Body(...)):
    endpoint = (payload.get("endpoint") or "").strip()
    with db() as conn:
        conn.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))
    return {"ok": True}


@router.get("/api/push/vapid-key")
def api_push_vapid_key():
    return {"ok": True, "key": cfg("vapid_public_key", "")}


@router.post("/api/push/generate-vapid")
def api_generate_vapid():
    """Genera un par de claves VAPID con openssl y las persiste en settings."""
    try:
        with tempfile.TemporaryDirectory() as tmp:
            key_path = os.path.join(tmp, "vapid.pem")
            subprocess.run(
                ["openssl", "ecparam", "-name", "prime256v1", "-genkey", "-noout", "-out", key_path],
                check=True, capture_output=True,
            )
            priv_der = subprocess.run(
                ["openssl", "ec", "-in", key_path, "-outform", "DER"],
                check=True, capture_output=True,
            ).stdout
            pub_der = subprocess.run(
                ["openssl", "ec", "-in", key_path, "-pubout", "-outform", "DER"],
                check=True, capture_output=True,
            ).stdout

        def b64u(b: bytes) -> str:
            return base64.urlsafe_b64encode(b).rstrip(b"=").decode()

        pub_key  = b64u(pub_der[-65:])
        priv_key = b64u(priv_der[-32:])

        with db() as conn:
            conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('vapid_public_key',?)",  (pub_key,))
            conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('vapid_private_key',?)", (priv_key,))

        from config import _cfg
        _cfg["vapid_public_key"]  = pub_key
        _cfg["vapid_private_key"] = priv_key

        return {"ok": True, "public_key": pub_key}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


# ═══════════════════════════════════════════════════════════════
#  TLS
# ═══════════════════════════════════════════════════════════════

@router.get("/api/tls/ca.crt")
def download_ca():
    ca_path = "/data/certs/ca.crt"
    if not os.path.exists(ca_path):
        return JSONResponse(
            {"error": "Certificado CA no generado aún. Espera al primer arranque."},
            status_code=404,
        )
    return FileResponse(
        ca_path, media_type="application/x-x509-ca-cert",
        headers={"Content-Disposition": 'attachment; filename="AuditorIPs-CA.crt"'},
    )


@router.get("/tls-info", response_class=HTMLResponse)
def tls_info(request: Request):
    host = request.headers.get("host", "192.168.1.x:8088")
    base = f"https://{host}"
    html = f"""<!doctype html><html><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>Instalar certificado — Auditor IPs</title>
<style>
  body{{font-family:sans-serif;max-width:500px;margin:40px auto;padding:20px;background:#1a1a2e;color:#eee}}
  h2{{color:#4dffb5}}
  a.btn{{display:block;background:#4dffb5;color:#000;padding:14px;border-radius:10px;
         text-align:center;text-decoration:none;font-weight:700;font-size:1.1rem;margin:20px 0}}
  ol{{line-height:2.2rem}}
  code{{background:rgba(255,255,255,.1);padding:2px 6px;border-radius:4px}}
</style></head><body>
<h2>📱 Instalar certificado de seguridad</h2>
<p>Para eliminar el aviso <strong>"No seguro"</strong> en tu móvil Android:</p>
<a class=btn href="{base}/api/tls/ca.crt">⬇️ Descargar certificado CA</a>
<ol>
  <li>Pulsa el botón de arriba para descargar <code>AuditorIPs-CA.crt</code></li>
  <li>Ve a <strong>Ajustes → Seguridad → Más ajustes → Instalar desde almacenamiento</strong></li>
  <li>Selecciona el fichero descargado</li>
  <li>Ponle un nombre (ej: <em>AuditorIPs CA</em>) y confirma</li>
  <li>Reinicia Chrome y accede a <code>{base}</code></li>
</ol>
<p style="opacity:.6;font-size:.85rem">El certificado es local y solo válido en tu red.</p>
<p><a href="/" style="color:#4dffb5">← Volver a Auditor IPs</a></p>
</body></html>"""
    return HTMLResponse(html)


# ═══════════════════════════════════════════════════════════════
#  Export XLSX
# ═══════════════════════════════════════════════════════════════

@router.get("/export.xlsx")
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"ok": False, "error": "openpyxl no instalado"}, status_code=500)

    app_tz = get_app_tz(cfg("app_tz", "Europe/Madrid"))

    with db() as conn:
        host_rows = conn.execute("""
            SELECT h.ip, h.mac, h.nmap_hostname, h.dns_name, h.manual_name, h.notes,
                   COALESCE(t.name,'') AS type_name,
                   h.first_seen, h.last_seen, h.last_change, h.status
            FROM hosts h LEFT JOIN host_types t ON t.id = h.type_id
            ORDER BY h.status DESC, h.last_seen DESC
        """).fetchall()
        scan_rows = conn.execute("""
            SELECT id, started_at, finished_at, cidr, online_hosts, offline_hosts,
                   new_hosts, events_sent, discord_sent, discord_error
            FROM scans ORDER BY id DESC LIMIT 500
        """).fetchall()
        uptime_rows = conn.execute("""
            SELECT ip, date, online_seconds, offline_seconds
            FROM host_uptime ORDER BY ip ASC, date ASC
        """).fetchall()
        svc_rows = conn.execute("""
            SELECT s.name, s.host, s.port, s.protocol, s.service_type,
                   sc.status, sc.latency_ms, sc.checked_at, sc.error
            FROM services s
            LEFT JOIN service_checks sc ON sc.id = (
                SELECT id FROM service_checks WHERE service_id=s.id ORDER BY checked_at DESC LIMIT 1
            )
            ORDER BY s.name
        """).fetchall()

    wb = Workbook()

    # ── Hosts ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Hosts"
    hdr = ["IP", "MAC", "Hostname", "DNS", "Nombre manual", "Tipo", "Notas",
           "Primera vez", "Última vez", "Último cambio", "Visto hace", "Estado"]
    ws.append(hdr)
    for r in host_rows:
        ws.append([
            r["ip"], r["mac"] or "", r["nmap_hostname"] or "", r["dns_name"] or "",
            r["manual_name"] or "", r["type_name"] or "", r["notes"] or "",
            to_local_str(r["first_seen"]), to_local_str(r["last_seen"]),
            to_local_str(r["last_change"]), human_since(r["last_seen"]),
            r["status"] or "",
        ])
    for col in range(1, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ── Ejecuciones ────────────────────────────────────────────
    ws2 = wb.create_sheet("Ejecuciones")
    hdr2 = ["ID", "Inicio", "Fin", "Rango", "Online", "Offline",
            "Nuevos", "Eventos", "Discord", "Error Discord"]
    ws2.append(hdr2)
    for r in scan_rows:
        ws2.append([
            r["id"], to_local_str(r["started_at"]), to_local_str(r["finished_at"]),
            r["cidr"] or "", r["online_hosts"] or 0, r["offline_hosts"] or 0,
            r["new_hosts"] or 0, r["events_sent"] or 0,
            "SI" if r["discord_sent"] else "NO", r["discord_error"] or "",
        ])
    for col in range(1, len(hdr2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ── Uptime ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Uptime")
    hdr3 = ["IP", "Fecha", "Online (h)", "Offline (h)", "Uptime %"]
    ws3.append(hdr3)
    for r in uptime_rows:
        total = r["online_seconds"] + r["offline_seconds"]
        pct   = round(r["online_seconds"] * 100 / total, 1) if total > 0 else ""
        ws3.append([
            r["ip"], r["date"],
            round(r["online_seconds"]  / 3600, 2),
            round(r["offline_seconds"] / 3600, 2),
            pct,
        ])
    for col in range(1, len(hdr3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    # ── Servicios ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Servicios")
    hdr4 = ["Nombre", "Host", "Puerto", "Protocolo", "Tipo",
            "Estado", "Latencia (ms)", "Último check", "Error"]
    ws4.append(hdr4)
    for r in svc_rows:
        ws4.append([
            r["name"], r["host"], r["port"], r["protocol"], r["service_type"] or "",
            r["status"] or "", r["latency_ms"] or "",
            to_local_str(r["checked_at"]) if r["checked_at"] else "",
            r["error"] or "",
        ])
    for col in range(1, len(hdr4) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"auditor_ips_{datetime.now(app_tz).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
#  Exportación programada a Excel
# ═══════════════════════════════════════════════════════════════

def _build_hosts_xlsx(output_dir: str) -> str:
    """
    Genera hosts_network_data--YYYY-MM-DD.xlsx en output_dir.
    Devuelve la ruta completa del fichero generado.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter as gcl
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no instalado")

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"hosts_network_data--{date_str}.xlsx"
    filepath = str(Path(output_dir) / filename)

    with db() as conn:
        # Columnas opcionales: se leen con COALESCE para no fallar si no existen en la BD
        col_info = {row[1] for row in conn.execute("PRAGMA table_info(hosts)").fetchall()}
        _opt = lambda col: f"h.{col}" if col in col_info else f"'' AS {col}"
        query = f"""
            SELECT h.ip, h.mac,
                   {_opt('vendor')},
                   h.nmap_hostname AS hostname, h.manual_name, h.dns_name,
                   h.status,
                   {_opt('known')},
                   COALESCE(t.name,'') AS type_name,
                   h.notes, h.first_seen, h.last_seen,
                   {_opt('open_ports')},
                   {_opt('router_hostname')},
                   {_opt('router_lease_type')},
                   {_opt('tags')}
            FROM hosts h LEFT JOIN host_types t ON t.id = h.type_id
            ORDER BY
                CASE WHEN h.status='online'        THEN 0
                     WHEN h.status='online_silent' THEN 1
                     ELSE 2 END,
                h.ip
        """
        hosts = conn.execute(query).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hosts"
    ws.sheet_view.showGridLines = False

    # Estilos
    HEADER_FILL = PatternFill("solid", fgColor="0F172A")
    ONLINE_FILL = PatternFill("solid", fgColor="14532D")
    SILENT_FILL = PatternFill("solid", fgColor="713F12")
    OFFLIN_FILL = PatternFill("solid", fgColor="3B0764")
    DEFLT_FILL  = PatternFill("solid", fgColor="1E293B")
    WHITE_FONT  = Font(color="FFFFFF", name="Calibri", size=10)
    HEADER_FONT = Font(color="38BDF8", bold=True, name="Calibri", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    THIN   = Side(style="thin", color="334155")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    COLUMNS = [
        ("IP", 14), ("MAC", 18), ("Fabricante", 22), ("Hostname", 22),
        ("Nombre manual", 22), ("DNS", 22), ("Estado", 14), ("Conocido", 9),
        ("Tipo", 16), ("Notas", 30), ("Primera vez", 18), ("Última vez", 18),
        ("Puertos", 28), ("Hostname DHCP", 20), ("Lease tipo", 12), ("Etiquetas", 24),
    ]

    # Fila título
    ws.merge_cells(f"A1:{gcl(len(COLUMNS))}1")
    c = ws["A1"]
    c.value     = f"Auditor IPs — Inventario de hosts  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font      = Font(color="38BDF8", bold=True, name="Calibri", size=13)
    c.fill      = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Cabecera
    for ci, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[gcl(ci)].width = width
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    STATUS_FILLS = {"online": ONLINE_FILL, "online_silent": SILENT_FILL, "offline": OFFLIN_FILL}
    STATUS_LABELS = {
        "online": "Online", "online_silent": "Silent",
        "offline": "Offline", "unknown": "Unknown",
    }

    for ri, h in enumerate(hosts, 3):
        status   = h["status"] or "unknown"
        row_fill = STATUS_FILLS.get(status, DEFLT_FILL)
        values   = [
            h["ip"], h["mac"] or "", h["vendor"] or "",
            h["hostname"] or "", h["manual_name"] or "", h["dns_name"] or "",
            STATUS_LABELS.get(status, status),
            "Sí" if h["known"] else "No",
            h["type_name"] or "", h["notes"] or "",
            to_local_str(h["first_seen"]) if h["first_seen"] else "",
            to_local_str(h["last_seen"])  if h["last_seen"]  else "",
            h["open_ports"] or "", h["router_hostname"] or "",
            h["router_lease_type"] or "", h["tags"] or "",
        ]
        WIDE_COLS = {4, 5, 6, 10, 13, 16}
        for ci, value in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = row_fill; cell.font = WHITE_FONT
            cell.alignment = LEFT if ci in WIDE_COLS else CENTER
            cell.border = BORDER
        ws.row_dimensions[ri].height = 15

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False
    total   = len(hosts)
    online  = sum(1 for h in hosts if h["status"] == "online")
    silent  = sum(1 for h in hosts if h["status"] == "online_silent")
    offline = sum(1 for h in hosts if h["status"] == "offline")
    known   = sum(1 for h in hosts if h["known"])
    for ri, (k, v) in enumerate([
        ("Generado el",  datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Total hosts",  total),   ("Online",       online),
        ("Silent",       silent),  ("Offline",      offline),
        ("Conocidos",    known),   ("Desconocidos", total - known),
    ], 1):
        ck = ws2.cell(row=ri, column=1, value=k)
        cv = ws2.cell(row=ri, column=2, value=v)
        ck.font = Font(color="38BDF8", bold=True, name="Calibri", size=10)
        cv.font = Font(color="FFFFFF", name="Calibri", size=10)
        ck.fill = cv.fill = HEADER_FILL
        ck.alignment = cv.alignment = LEFT
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 22
    wb.active = ws

    wb.save(filepath)

    with db() as conn:
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                     ("export_xlsx_last_run", datetime.now().isoformat()))
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                     ("export_xlsx_last_file", filename))

    print(f"[export] Generado: {filepath} ({len(hosts)} hosts)")
    return filepath


def _reschedule_export():
    """Registra o actualiza el job de exportación según la configuración actual."""
    sched = _get_scheduler()
    if sched is None:
        return

    JOB_ID = "export_xlsx_job"
    try:
        sched.remove_job(JOB_ID)
    except Exception:
        pass

    with db() as conn:
        def _s(k, d=""):
            r = conn.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone()
            return r["value"] if r else d
        enabled   = _s("export_xlsx_enabled",  "0") == "1"
        out_path  = _s("export_xlsx_path",      "/data/exports")
        frequency = _s("export_xlsx_frequency", "weekly")
        day       = int(_s("export_xlsx_day",   "0"))
        hour      = int(_s("export_xlsx_hour",  "6"))

    if not enabled:
        return

    def _job():
        try:
            _build_hosts_xlsx(out_path)
        except Exception as e:
            print(f"[export] Error en exportación programada: {e}")

    DIAS = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
    if frequency == "daily":
        sched.add_job(_job, "cron", hour=hour, minute=0,
                      id=JOB_ID, replace_existing=True)
        print(f"[export] Job diario a las {hour:02d}:00")
    elif frequency == "weekly":
        sched.add_job(_job, "cron", day_of_week=day, hour=hour, minute=0,
                      id=JOB_ID, replace_existing=True)
        print(f"[export] Job semanal: {DIAS[day]} a las {hour:02d}:00")
    elif frequency == "monthly":
        sched.add_job(_job, "cron", day=day, hour=hour, minute=0,
                      id=JOB_ID, replace_existing=True)
        print(f"[export] Job mensual: día {day} a las {hour:02d}:00")


# Alias público por si main.py lo llama directamente
reschedule_export = _reschedule_export


@router.get("/api/export/xlsx/config")
def get_export_config():
    with db() as conn:
        def _s(k, d=""):
            r = conn.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone()
            return r["value"] if r else d
        return {
            "enabled":   _s("export_xlsx_enabled",  "0") == "1",
            "path":      _s("export_xlsx_path",      "/data/exports"),
            "frequency": _s("export_xlsx_frequency", "weekly"),
            "day":       int(_s("export_xlsx_day",   "0")),
            "hour":      int(_s("export_xlsx_hour",  "6")),
            "last_run":  _s("export_xlsx_last_run",  ""),
            "last_file": _s("export_xlsx_last_file", ""),
        }


@router.put("/api/export/xlsx/config")
def save_export_config(payload: Dict[str, Any] = Body(...)):
    freq = payload.get("frequency", "weekly")
    if freq not in {"daily", "weekly", "monthly"}:
        return JSONResponse(status_code=400, content={"error": "Frecuencia inválida"})
    hour = int(payload.get("hour", 6))
    day  = int(payload.get("day",  0))
    if not (0 <= hour <= 23):
        return JSONResponse(status_code=400, content={"error": "Hora fuera de rango (0-23)"})

    with db() as conn:
        for k, v in [
            ("export_xlsx_enabled",   "1" if payload.get("enabled") else "0"),
            ("export_xlsx_path",      str(payload.get("path", "/data/exports")).strip()),
            ("export_xlsx_frequency", freq),
            ("export_xlsx_day",       str(day)),
            ("export_xlsx_hour",      str(hour)),
        ]:
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, v))

    _reschedule_export()
    return {"ok": True}


@router.post("/api/export/xlsx/now")
def export_xlsx_now():
    with db() as conn:
        r = conn.execute("SELECT value FROM settings WHERE key='export_xlsx_path'").fetchone()
        out_path = r["value"] if r else "/data/exports"
    try:
        filepath = _build_hosts_xlsx(out_path)
        return {"ok": True, "file": Path(filepath).name, "path": filepath}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
#  Config Procesos — Scripts monitorizados (S17)
# ═══════════════════════════════════════════════════════════════

SCRIPTS_STATUS_DIR: str = os.getenv("SCRIPTS_STATUS_DIR", "/data/scripts_status")


@router.get("/api/config/scripts/available")
def api_scripts_available():
    """Lista los .status.json disponibles en el volumen (para autocompletar al añadir)."""
    try:
        if not os.path.isdir(SCRIPTS_STATUS_DIR):
            return {"ok": True, "files": []}
        files = sorted([
            f.replace(".status.json", "")
            for f in os.listdir(SCRIPTS_STATUS_DIR)
            if f.endswith(".status.json")
        ])
        return {"ok": True, "files": files}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@router.get("/api/config/scripts")
def api_scripts_list():
    """Lista los scripts monitorizados configurados en BD."""
    with db() as conn:
        rows = conn.execute(
            "SELECT id, script_name, label, description, color, active, sort_order, created_at "
            "FROM monitored_scripts ORDER BY sort_order ASC, id ASC"
        ).fetchall()
    return {"ok": True, "scripts": [dict(r) for r in rows]}


@router.post("/api/config/scripts")
def api_scripts_create(payload: Dict[str, Any] = Body(...)):
    """Añade un nuevo script monitorizado."""
    script_name = (payload.get("script_name") or "").strip()
    if not script_name:
        return JSONResponse(status_code=400, content={"ok": False, "error": "script_name requerido"})
    label       = (payload.get("label") or "").strip()
    description = (payload.get("description") or "").strip()
    color       = (payload.get("color") or "").strip()
    active      = int(payload.get("active", 1))
    sort_order  = int(payload.get("sort_order", 0))
    try:
        with db() as conn:
            conn.execute(
                "INSERT INTO monitored_scripts "
                "(script_name, label, description, color, active, sort_order, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (script_name, label, description, color, active, sort_order, utc_now_iso())
            )
            new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"ok": True, "id": new_id}
    except Exception as e:
        if "UNIQUE" in str(e):
            return JSONResponse(status_code=409, content={"ok": False, "error": f"'{script_name}' ya existe"})
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@router.put("/api/config/scripts/{script_id}")
def api_scripts_update(script_id: int, payload: Dict[str, Any] = Body(...)):
    """Actualiza un script monitorizado."""
    with db() as conn:
        row = conn.execute("SELECT id FROM monitored_scripts WHERE id=?", (script_id,)).fetchone()
        if not row:
            return JSONResponse(status_code=404, content={"ok": False, "error": "No encontrado"})
        fields, vals = [], []
        for col in ("script_name", "label", "description", "color", "sort_order"):
            if col in payload:
                fields.append(f"{col}=?")
                vals.append((payload[col] or "").strip() if isinstance(payload[col], str) else payload[col])
        if "active" in payload:
            fields.append("active=?")
            vals.append(int(payload["active"]))
        if not fields:
            return {"ok": True, "updated": 0}
        vals.append(script_id)
        conn.execute(f"UPDATE monitored_scripts SET {', '.join(fields)} WHERE id=?", vals)
    return {"ok": True}


@router.delete("/api/config/scripts/{script_id}")
def api_scripts_delete(script_id: int):
    """Elimina un script monitorizado."""
    with db() as conn:
        cur = conn.execute("DELETE FROM monitored_scripts WHERE id=?", (script_id,))
    if cur.rowcount == 0:
        return JSONResponse(status_code=404, content={"ok": False, "error": "No encontrado"})
    return {"ok": True}


@router.post("/api/config/scripts/reorder")
def api_scripts_reorder(payload: Dict[str, Any] = Body(...)):
    """Actualiza sort_order de varios scripts de una vez. payload: {ids: [1,3,2,...]}"""
    ids = payload.get("ids") or []
    with db() as conn:
        for order, sid in enumerate(ids):
            conn.execute("UPDATE monitored_scripts SET sort_order=? WHERE id=?", (order, sid))
    return {"ok": True}


@router.post("/api/config/scripts/import-all")
def api_scripts_import_all():
    """
    Importa todos los .status.json del volumen a monitored_scripts (INSERT OR IGNORE).
    Devuelve cuántos se añadieron y cuántos ya existían.
    """
    try:
        if not os.path.isdir(SCRIPTS_STATUS_DIR):
            return JSONResponse(status_code=404, content={"ok": False, "error": "Directorio no encontrado"})
        files = sorted([
            f.replace(".status.json", "")
            for f in os.listdir(SCRIPTS_STATUS_DIR)
            if f.endswith(".status.json")
        ])
        added, skipped = 0, 0
        with db() as conn:
            existing = {r[0] for r in conn.execute("SELECT script_name FROM monitored_scripts").fetchall()}
            for i, name in enumerate(files):
                if name in existing:
                    skipped += 1
                    continue
                label = name.replace("_", " ").title()
                conn.execute(
                    "INSERT OR IGNORE INTO monitored_scripts "
                    "(script_name, label, description, color, active, sort_order, created_at) "
                    "VALUES (?, ?, ?, ?, 1, ?, ?)",
                    (name, label, "", "", i, utc_now_iso())
                )
                added += 1
        return {"ok": True, "added": added, "skipped": skipped, "total": len(files)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


# ═══════════════════════════════════════════════════════════════
#  Redes secundarias (Config → Redes — Sesión 21)
# ═══════════════════════════════════════════════════════════════

@router.get("/api/config/networks")
def api_networks_list():
    """Lista las redes secundarias configuradas."""
    with db() as conn:
        rows = conn.execute(
            "SELECT id, label, cidr, interface, enabled, created_at "
            "FROM secondary_networks ORDER BY id ASC"
        ).fetchall()
    return {"ok": True, "networks": [dict(r) for r in rows]}


@router.post("/api/config/networks")
def api_networks_create(payload: Dict[str, Any] = Body(...)):
    """Añade una red secundaria."""
    cidr      = (payload.get("cidr") or "").strip()
    label     = (payload.get("label") or "").strip()
    interface = (payload.get("interface") or "").strip()
    enabled   = int(payload.get("enabled", 1))
    if not cidr:
        return JSONResponse(status_code=400, content={"ok": False, "error": "cidr requerido"})
    try:
        with db() as conn:
            conn.execute(
                "INSERT INTO secondary_networks (label, cidr, interface, enabled, created_at) "
                "VALUES (?, ?, ?, ?, ?)",
                (label, cidr, interface, enabled, utc_now_iso())
            )
            new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"ok": True, "id": new_id}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@router.put("/api/config/networks/{net_id}")
def api_networks_update(net_id: int, payload: Dict[str, Any] = Body(...)):
    """Actualiza una red secundaria."""
    with db() as conn:
        row = conn.execute("SELECT id FROM secondary_networks WHERE id=?", (net_id,)).fetchone()
        if not row:
            return JSONResponse(status_code=404, content={"ok": False, "error": "No encontrado"})
        fields, vals = [], []
        for col in ("label", "cidr", "interface"):
            if col in payload:
                fields.append(f"{col}=?")
                vals.append((payload[col] or "").strip())
        if "enabled" in payload:
            fields.append("enabled=?")
            vals.append(int(payload["enabled"]))
        if not fields:
            return {"ok": True, "updated": 0}
        vals.append(net_id)
        conn.execute(f"UPDATE secondary_networks SET {', '.join(fields)} WHERE id=?", vals)
    return {"ok": True}


@router.delete("/api/config/networks/{net_id}")
def api_networks_delete(net_id: int):
    """Elimina una red secundaria."""
    with db() as conn:
        cur = conn.execute("DELETE FROM secondary_networks WHERE id=?", (net_id,))
    if cur.rowcount == 0:
        return JSONResponse(status_code=404, content={"ok": False, "error": "No encontrado"})
    return {"ok": True}


@router.get("/api/config/network/interfaces")
def api_network_interfaces():
    """
    Detecta las interfaces de red activas del host con su IP y CIDR.
    Usa 'ip -j addr' (formato JSON del sistema) disponible en cualquier Linux moderno.
    """
    import subprocess, json as _json
    try:
        out = subprocess.run(
            ["ip", "-j", "addr"],
            capture_output=True, text=True, timeout=5
        )
        ifaces = _json.loads(out.stdout)
        result = []
        for iface in ifaces:
            name = iface.get("ifname", "")
            if not name or name == "lo" or name.startswith(("docker", "br-", "veth", "virbr")):
                continue
            operstate = (iface.get("operstate") or "").upper()
            flags = iface.get("flags", [])
            is_up = operstate == "UP" or "UP" in flags
            addrs = []
            for addr in iface.get("addr_info", []):
                if addr.get("family") == "inet":   # solo IPv4
                    addrs.append(f"{addr['local']}/{addr['prefixlen']}")
            if not is_up and not addrs:
                continue
            result.append({
                "name":    name,
                "state":   iface.get("operstate", "UNKNOWN"),
                "mac":     iface.get("address", ""),
                "addrs":   addrs,
            })
        return {"ok": True, "interfaces": result}
    except Exception as e:
        return {"ok": False, "interfaces": [], "error": str(e)}


# ═══════════════════════════════════════════════════════════════
#  Archivos estáticos PWA
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
#  Discrepancias nmap / router
# ═══════════════════════════════════════════════════════════════

@router.get("/api/scan/discrepancies")
def api_discrepancies_list():
    """Lista de IPs vistas por nmap pero no por el router."""
    try:
        with db() as conn:
            rows = conn.execute("""
                SELECT d.id, d.ip, d.mac, d.first_seen, d.last_seen, d.times_seen,
                       d.accepted, d.accepted_at, d.note,
                       h.manual_name, h.nmap_hostname, h.router_hostname, h.vendor
                FROM scan_discrepancies d
                LEFT JOIN hosts h ON h.ip = d.ip
                ORDER BY d.accepted ASC, d.times_seen DESC
            """).fetchall()
        return {"ok": True, "discrepancies": [dict(r) for r in rows]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.post("/api/scan/discrepancies/{disc_id}/accept")
def api_discrepancy_accept(disc_id: int, payload: Dict[str, Any] = Body(default={})):
    """Marca una discrepancia como aceptada."""
    note = (payload.get("note") or "").strip()
    try:
        with db() as conn:
            conn.execute("""
                UPDATE scan_discrepancies
                SET accepted=1, accepted_at=?, note=?
                WHERE id=?
            """, (utc_now_iso(), note or None, disc_id))
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.delete("/api/scan/discrepancies/{disc_id}")
def api_discrepancy_delete(disc_id: int):
    """Elimina una discrepancia."""
    try:
        with db() as conn:
            conn.execute("DELETE FROM scan_discrepancies WHERE id=?", (disc_id,))
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.post("/api/scan/discrepancies/accept-all")
def api_discrepancy_accept_all():
    """Acepta todas las discrepancias pendientes."""
    try:
        now = utc_now_iso()
        with db() as conn:
            conn.execute("UPDATE scan_discrepancies SET accepted=1, accepted_at=? WHERE accepted=0", (now,))
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@router.get("/manifest.json")
def serve_manifest():
    return FileResponse(
        "manifest.json",
        media_type="application/manifest+json",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


@router.get("/sw.js")
def serve_sw():
    return FileResponse(
        "sw.js",
        media_type="application/javascript",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


# ══════════════════════════════════════════════════════════════
#  Exportación histórica por rango de fechas (Sesión 24)
# ══════════════════════════════════════════════════════════════

@router.get("/api/export/history")
def api_export_history(date_from: str = "", date_to: str = ""):
    """
    Genera y descarga un Excel con tres hojas:
      - Uptime        (host_uptime por día)
      - Latencia      (host_latency, muestra reducida por host/día)
      - Servicios     (service_checks)
    Parámetros: date_from y date_to en formato YYYY-MM-DD.
    Si se omiten, devuelve los últimos 30 días.
    """
    from datetime import datetime, timedelta
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "openpyxl no instalado"}, status_code=500)

    # ── Calcular rango ────────────────────────────────────────
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d") if date_from else datetime.utcnow() - timedelta(days=30)
        dt_to   = datetime.strptime(date_to,   "%Y-%m-%d") if date_to   else datetime.utcnow()
    except ValueError:
        return JSONResponse({"error": "Formato de fecha inválido. Usa YYYY-MM-DD"}, status_code=400)

    str_from = dt_from.strftime("%Y-%m-%d")
    str_to   = dt_to.strftime("%Y-%m-%d")
    # Para latencia y servicios (ISO timestamps) necesitamos el límite superior al final del día
    iso_to   = dt_to.strftime("%Y-%m-%d") + "T23:59:59"
    iso_from = dt_from.strftime("%Y-%m-%d") + "T00:00:00"

    # ── Estilos comunes ────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    ROW_FILL  = PatternFill("solid", fgColor="1E293B")
    ALT_FILL  = PatternFill("solid", fgColor="172033")
    HDR_FONT  = Font(color="38BDF8", bold=True, name="Calibri", size=10)
    ROW_FONT  = Font(color="FFFFFF", name="Calibri", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    THIN      = Side(style="thin", color="334155")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _title_row(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.font = Font(color="38BDF8", bold=True, name="Calibri", size=12)
        c.fill = HDR_FILL
        c.alignment = CENTER
        ws.row_dimensions[1].height = 22

    def _header_row(ws, cols, row=2):
        for i, (label, width) in enumerate(cols, start=1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = BORDER
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[row].height = 16
        ws.freeze_panes = ws.cell(row=3, column=1)

    def _data_row(ws, row_idx, values, aligns=None):
        fill = ROW_FILL if row_idx % 2 == 1 else ALT_FILL
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=i, value=v)
            c.font = ROW_FONT; c.fill = fill
            c.alignment = (aligns[i-1] if aligns else CENTER)
            c.border = BORDER
        ws.row_dimensions[row_idx].height = 14

    wb = Workbook()

    # ══ Hoja 1: Uptime ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Uptime"
    ws1.sheet_view.showGridLines = False

    with db() as conn:
        uptime_rows = conn.execute("""
            SELECT u.ip, COALESCE(h.manual_name, h.hostname, u.ip) as name,
                   u.date, u.online_seconds, u.offline_seconds
            FROM host_uptime u
            LEFT JOIN hosts h ON h.ip = u.ip
            WHERE u.date BETWEEN ? AND ?
            ORDER BY u.ip, u.date
        """, (str_from, str_to)).fetchall()

    UCOLS = [("IP",14),("Nombre",22),("Fecha",12),("Online (s)",12),("Offline (s)",12),
             ("Online %",10),("Online (h:m)",12)]
    _title_row(ws1, f"Uptime  ·  {str_from} → {str_to}  ·  {len(uptime_rows)} registros", len(UCOLS))
    _header_row(ws1, UCOLS)
    for idx, r in enumerate(uptime_rows, start=1):
        total = (r["online_seconds"] or 0) + (r["offline_seconds"] or 0)
        pct   = round(r["online_seconds"] * 100 / total, 1) if total > 0 else 0
        secs  = r["online_seconds"] or 0
        hm    = f"{secs//3600}h {(secs%3600)//60}m"
        _data_row(ws1, idx+2,
                  [r["ip"], r["name"], r["date"],
                   r["online_seconds"], r["offline_seconds"], pct, hm],
                  [CENTER,LEFT,CENTER,CENTER,CENTER,CENTER,CENTER])

    # ══ Hoja 2: Latencia ══════════════════════════════════════
    ws2 = wb.create_sheet("Latencia")
    ws2.sheet_view.showGridLines = False

    with db() as conn:
        # Muestra reducida: promedio/min/max por host y hora para no generar ficheros enormes
        lat_rows = conn.execute("""
            SELECT l.ip, COALESCE(h.manual_name, h.hostname, l.ip) as name,
                   strftime('%Y-%m-%d %H:00', l.scanned_at) as hour_bucket,
                   COUNT(*) as samples,
                   ROUND(AVG(l.latency_ms),1) as avg_ms,
                   ROUND(MIN(l.latency_ms),1) as min_ms,
                   ROUND(MAX(l.latency_ms),1) as max_ms
            FROM host_latency l
            LEFT JOIN hosts h ON h.ip = l.ip
            WHERE l.scanned_at BETWEEN ? AND ?
              AND l.latency_ms IS NOT NULL
            GROUP BY l.ip, hour_bucket
            ORDER BY l.ip, hour_bucket
        """, (iso_from, iso_to)).fetchall()

    LCOLS = [("IP",14),("Nombre",22),("Hora",18),("Muestras",10),
             ("Avg ms",10),("Min ms",10),("Max ms",10)]
    _title_row(ws2, f"Latencia (agrupada por hora)  ·  {str_from} → {str_to}  ·  {len(lat_rows)} registros", len(LCOLS))
    _header_row(ws2, LCOLS)
    for idx, r in enumerate(lat_rows, start=1):
        _data_row(ws2, idx+2,
                  [r["ip"], r["name"], r["hour_bucket"],
                   r["samples"], r["avg_ms"], r["min_ms"], r["max_ms"]],
                  [CENTER,LEFT,CENTER,CENTER,CENTER,CENTER,CENTER])

    # ══ Hoja 3: Servicios ═════════════════════════════════════
    ws3 = wb.create_sheet("Servicios")
    ws3.sheet_view.showGridLines = False

    with db() as conn:
        svc_rows = conn.execute("""
            SELECT s.name as svc_name, s.host, s.port, s.protocol,
                   sc.checked_at, sc.status, sc.latency_ms, sc.error
            FROM service_checks sc
            JOIN services s ON s.id = sc.service_id
            WHERE sc.checked_at BETWEEN ? AND ?
            ORDER BY s.name, sc.checked_at DESC
        """, (iso_from, iso_to)).fetchall()

    SCOLS = [("Servicio",20),("Host",18),("Puerto",8),("Protocolo",10),
             ("Comprobado",18),("Estado",10),("Latencia ms",12),("Error",40)]
    _title_row(ws3, f"Servicios  ·  {str_from} → {str_to}  ·  {len(svc_rows)} registros", len(SCOLS))
    _header_row(ws3, SCOLS)
    for idx, r in enumerate(svc_rows, start=1):
        _data_row(ws3, idx+2,
                  [r["svc_name"], r["host"], r["port"], r["protocol"],
                   r["checked_at"], r["status"], r["latency_ms"], r["error"] or ""],
                  [LEFT,LEFT,CENTER,CENTER,CENTER,CENTER,CENTER,LEFT])

    # ══ Generar en memoria y devolver como descarga ════════════
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"auditor_historico_{str_from}_{str_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
