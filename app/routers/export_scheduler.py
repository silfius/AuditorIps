# ============================================================
# PARCHE: Exportación programada de hosts a Excel
# Añadir estas funciones y rutas a app/routers/config_api.py
# ============================================================
#
# DEPENDENCIAS (ya presentes en el proyecto):
#   openpyxl, apscheduler, fastapi
#
# SETTINGS nuevos en la BD (se crean automáticamente):
#   export_xlsx_enabled    → "0" / "1"
#   export_xlsx_path       → ruta directorio en el host (ej: /data/exports)
#   export_xlsx_frequency  → "daily" / "weekly" / "monthly"
#   export_xlsx_day        → día semana 0-6 (lunes=0) o día mes 1-31
#   export_xlsx_hour       → hora de ejecución 0-23
#   export_xlsx_last_run   → ISO timestamp última ejecución (solo lectura)
#   export_xlsx_last_file  → nombre del último fichero generado
#
# INTEGRACIÓN EN config_api.py:
#   1. Añadir las importaciones del bloque "IMPORTS ADICIONALES"
#   2. Pegar las funciones y rutas al final del fichero
#   3. En set_scheduler() añadir:  _reschedule_export()
#
# INTEGRACIÓN EN main.py:
#   En startup(), después de r_config_api.set_scheduler(scheduler):
#     from routers.config_api import reschedule_export
#     reschedule_export()
# ============================================================

# ── IMPORTS ADICIONALES (añadir a los imports existentes) ─────────────────────
import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── VARIABLES DE MÓDULO (añadir junto a las existentes) ──────────────────────
_export_scheduler = None   # se inyecta desde set_scheduler()

# ── FUNCIÓN: inyección del scheduler ─────────────────────────────────────────
# NOTA: set_scheduler() ya existe en config_api.py.
# Añadir al final del set_scheduler() existente:
#
#   global _export_scheduler
#   _export_scheduler = sched
#   _reschedule_export()
#
# Si prefieres no tocar set_scheduler(), llama a reschedule_export() desde main.py
# después del scheduler.start()

# ── FUNCIÓN CORE: generar el fichero Excel ────────────────────────────────────

def _build_hosts_xlsx(output_dir: str) -> str:
    """
    Genera hosts_network_data--YYYY-MM-DD.xlsx en output_dir.
    Devuelve la ruta completa del fichero generado.
    """
    from database import db   # import relativo al contexto de config_api.py

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    date_str  = datetime.now().strftime("%Y-%m-%d")
    filename  = f"hosts_network_data--{date_str}.xlsx"
    filepath  = str(Path(output_dir) / filename)

    with db() as conn:
        hosts = conn.execute("""
            SELECT
                h.ip, h.mac, h.vendor,
                h.hostname, h.manual_name, h.dns_name,
                h.status, h.known,
                h.type_name,
                h.notes,
                h.first_seen, h.last_seen,
                h.open_ports,
                h.router_hostname, h.router_lease_type,
                h.tags
            FROM hosts h
            ORDER BY
                CASE WHEN h.status='online'  THEN 0
                     WHEN h.status='online_silent' THEN 1
                     ELSE 2 END,
                CAST(SUBSTR(h.ip, 1, INSTR(h.ip,'.')-1) AS INTEGER),
                CAST(SUBSTR(h.ip, INSTR(h.ip,'.')+1,
                     INSTR(SUBSTR(h.ip,INSTR(h.ip,'.')+1),'.')-1) AS INTEGER),
                CAST(SUBSTR(h.ip, LENGTH(h.ip)-INSTR(REVERSE(h.ip),'.')+2) AS INTEGER)
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hosts"

    # ── Estilos ───────────────────────────────────────────────────────────────
    DARK_FILL   = PatternFill("solid", fgColor="1E293B")
    HEADER_FILL = PatternFill("solid", fgColor="0F172A")
    ONLINE_FILL = PatternFill("solid", fgColor="14532D")
    SILENT_FILL = PatternFill("solid", fgColor="713F12")
    OFFLIN_FILL = PatternFill("solid", fgColor="3B0764")
    WHITE_FONT  = Font(color="FFFFFF", bold=False, name="Calibri", size=10)
    HEADER_FONT = Font(color="38BDF8", bold=True, name="Calibri", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    THIN        = Side(style="thin", color="334155")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"Auditor IPs — Inventario de hosts  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font  = Font(color="38BDF8", bold=True, name="Calibri", size=13)
    title_cell.fill  = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Cabecera ──────────────────────────────────────────────────────────────
    COLUMNS = [
        ("IP",            14),
        ("MAC",           18),
        ("Fabricante",    22),
        ("Hostname",      22),
        ("Nombre manual", 22),
        ("DNS",           22),
        ("Estado",        14),
        ("Conocido",       9),
        ("Tipo",          16),
        ("Notas",         30),
        ("Primera vez",   18),
        ("Última vez",    18),
        ("Puertos",       28),
        ("Hostname DHCP", 20),
        ("Lease tipo",    12),
        ("Etiquetas",     24),
    ]
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    # ── Datos ─────────────────────────────────────────────────────────────────
    STATUS_LABELS = {
        "online":        "🟢 Online",
        "online_silent": "🟡 Silent",
        "offline":       "🔴 Offline",
        "unknown":       "⚪ Unknown",
    }
    STATUS_FILLS = {
        "online":        ONLINE_FILL,
        "online_silent": SILENT_FILL,
        "offline":       OFFLIN_FILL,
    }

    for row_idx, h in enumerate(hosts, start=3):
        status    = h["status"] or "unknown"
        row_fill  = STATUS_FILLS.get(status, DARK_FILL)

        values = [
            h["ip"],
            h["mac"] or "",
            h["vendor"] or "",
            h["hostname"] or "",
            h["manual_name"] or "",
            h["dns_name"] or "",
            STATUS_LABELS.get(status, status),
            "Sí" if h["known"] else "No",
            h["type_name"] or "",
            h["notes"] or "",
            h["first_seen"] or "",
            h["last_seen"] or "",
            h["open_ports"] or "",
            h["router_hostname"] or "",
            h["router_lease_type"] or "",
            h["tags"] or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill       = row_fill
            cell.font       = WHITE_FONT
            cell.alignment  = LEFT if col_idx in (4,5,6,10,13,16) else CENTER
            cell.border     = BORDER
        ws.row_dimensions[row_idx].height = 15

    # ── Hoja resumen ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False
    total   = len(hosts)
    online  = sum(1 for h in hosts if h["status"] == "online")
    silent  = sum(1 for h in hosts if h["status"] == "online_silent")
    offline = sum(1 for h in hosts if h["status"] == "offline")
    known   = sum(1 for h in hosts if h["known"])

    resumen_data = [
        ("Generado el",      datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("",                 ""),
        ("Total hosts",      total),
        ("Online",           online),
        ("Online silent",    silent),
        ("Offline",          offline),
        ("Conocidos",        known),
        ("Desconocidos",     total - known),
    ]
    for r, (k, v) in enumerate(resumen_data, start=1):
        ck = ws2.cell(row=r, column=1, value=k)
        cv = ws2.cell(row=r, column=2, value=v)
        ck.font = Font(color="38BDF8", bold=True, name="Calibri", size=10)
        cv.font = Font(color="FFFFFF",             name="Calibri", size=10)
        ck.fill = cv.fill = HEADER_FILL
        ck.alignment = cv.alignment = LEFT
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 22

    ws2.sheet_view.tabSelected = False
    wb.active = ws

    wb.save(filepath)

    # Guardar metadata en settings
    from database import db as _db
    with _db() as conn:
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                     ("export_xlsx_last_run", datetime.now().isoformat()))
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",
                     ("export_xlsx_last_file", filename))

    print(f"[export] Excel generado: {filepath} ({len(hosts)} hosts)")
    return filepath


# ── SCHEDULER: registrar / reajustar el job ───────────────────────────────────

def _reschedule_export():
    """Registra o actualiza el job de exportación según la configuración actual."""
    from database import db as _db   # import relativo al contexto de config_api.py
    global _export_scheduler
    if _export_scheduler is None:
        return

    JOB_ID = "export_xlsx_job"

    # Leer config
    with _db() as conn:
        def _s(k, default=""):
            r = conn.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone()
            return r["value"] if r else default

        enabled   = _s("export_xlsx_enabled",   "0") == "1"
        out_path  = _s("export_xlsx_path",       "/data/exports")
        frequency = _s("export_xlsx_frequency",  "weekly")
        day       = int(_s("export_xlsx_day",    "0"))   # 0=lunes o 1 para mensual
        hour      = int(_s("export_xlsx_hour",   "6"))

    # Eliminar job anterior si existe
    try:
        _export_scheduler.remove_job(JOB_ID)
    except Exception:
        pass

    if not enabled:
        print("[export] Exportación programada desactivada")
        return

    def _job():
        try:
            _build_hosts_xlsx(out_path)
        except Exception as e:
            print(f"[export] Error en exportación programada: {e}")

    if frequency == "daily":
        _export_scheduler.add_job(
            _job, "cron",
            hour=hour, minute=0,
            id=JOB_ID, replace_existing=True,
        )
        print(f"[export] Job diario a las {hour:02d}:00")

    elif frequency == "weekly":
        _export_scheduler.add_job(
            _job, "cron",
            day_of_week=day, hour=hour, minute=0,
            id=JOB_ID, replace_existing=True,
        )
        DIAS = ["lun","mar","mié","jue","vie","sáb","dom"]
        print(f"[export] Job semanal: {DIAS[day]} a las {hour:02d}:00")

    elif frequency == "monthly":
        _export_scheduler.add_job(
            _job, "cron",
            day=day, hour=hour, minute=0,
            id=JOB_ID, replace_existing=True,
        )
        print(f"[export] Job mensual: día {day} a las {hour:02d}:00")


# Alias público para main.py
reschedule_export = _reschedule_export


# ── API ENDPOINTS (añadir al router existente de config_api.py) ───────────────

# from fastapi import APIRouter  ← ya existe
# router = ...                   ← ya existe

# GET /api/export/xlsx/config
# router.add_api_route("/api/export/xlsx/config", get_export_config, methods=["GET"])

async def get_export_config():
    """Devuelve la configuración de exportación programada + metadata última ejecución."""
    from database import db as _db
    with _db() as conn:
        def _s(k, d=""):
            r = conn.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone()
            return r["value"] if r else d
        return {
            "enabled":   _s("export_xlsx_enabled",   "0") == "1",
            "path":      _s("export_xlsx_path",       "/data/exports"),
            "frequency": _s("export_xlsx_frequency",  "weekly"),
            "day":       int(_s("export_xlsx_day",    "0")),
            "hour":      int(_s("export_xlsx_hour",   "6")),
            "last_run":  _s("export_xlsx_last_run",   ""),
            "last_file": _s("export_xlsx_last_file",  ""),
        }


# PUT /api/export/xlsx/config
# router.add_api_route("/api/export/xlsx/config", save_export_config, methods=["PUT"])

async def save_export_config(data: dict):
    """Guarda la configuración y reactiva el scheduler."""
    from database import db as _db
    from fastapi.responses import JSONResponse

    allowed_frequencies = {"daily", "weekly", "monthly"}
    freq = data.get("frequency", "weekly")
    if freq not in allowed_frequencies:
        return JSONResponse(status_code=400, content={"error": "Frecuencia inválida"})

    hour = int(data.get("hour", 6))
    day  = int(data.get("day",  0))
    if not (0 <= hour <= 23):
        return JSONResponse(status_code=400, content={"error": "Hora fuera de rango (0-23)"})

    with _db() as conn:
        pairs = [
            ("export_xlsx_enabled",   "1" if data.get("enabled") else "0"),
            ("export_xlsx_path",      str(data.get("path", "/data/exports")).strip()),
            ("export_xlsx_frequency", freq),
            ("export_xlsx_day",       str(day)),
            ("export_xlsx_hour",      str(hour)),
        ]
        for k, v in pairs:
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, v))

    _reschedule_export()
    return {"ok": True}


# POST /api/export/xlsx/now
# router.add_api_route("/api/export/xlsx/now", export_xlsx_now, methods=["POST"])

async def export_xlsx_now():
    """Dispara una exportación inmediata."""
    from database import db as _db
    from fastapi.responses import JSONResponse
    with _db() as conn:
        r = conn.execute("SELECT value FROM settings WHERE key='export_xlsx_path'").fetchone()
        out_path = r["value"] if r else "/data/exports"
    try:
        filepath = _build_hosts_xlsx(out_path)
        return {"ok": True, "file": Path(filepath).name, "path": filepath}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
